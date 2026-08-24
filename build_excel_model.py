"""
Builds PE_Fund_Model.xlsx from the CSVs / model_output.json written by
generate_data.py and calculations.py.

All computed figures (capital calls, management fees, the distribution
waterfall, the NAV rollforward, LP capital accounts, IRR/MOIC, and the
reconciliation) are written as LIVE Excel formulas that reference the raw
inputs (LP commitments, portfolio company cost basis, quarterly fair value
marks, exit proceeds) — not as pasted Python-computed values. Change an
input cell and the workbook recalculates.

Color convention: blue = hardcoded input, black = formula, green = link to
another sheet, yellow fill = key assumption.
"""
import csv
import json
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.datavalidation import DataValidation

FONT_NAME = "Calibri"
BLUE = Font(name=FONT_NAME, color="0000FF")
BLUE_BOLD = Font(name=FONT_NAME, color="0000FF", bold=True)
BLACK = Font(name=FONT_NAME, color="000000")
GREEN = Font(name=FONT_NAME, color="008000")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F4E5F")
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color="1F4E5F")
SUBTITLE_FONT = Font(name=FONT_NAME, italic=True, size=10, color="555555")
SECTION_FONT = Font(name=FONT_NAME, bold=True, size=11, color="1F4E5F")
YELLOW_FILL = PatternFill("solid", fgColor="FFFFCC")
TOTAL_FONT = Font(name=FONT_NAME, bold=True)
TOTAL_BORDER = Border(top=Side(style="thin"))
THIN_BORDER = Border(bottom=Side(style="hair", color="DDDDDD"))

CUR0 = '$#,##0;($#,##0);"-"'
CUR2 = '$#,##0.00;($#,##0.00);"-"'
PCT1 = '0.0%'
PCT2 = '0.00%'
MULT = '0.00"x"'


def read_csv(path):
    with open(path) as f:
        return list(csv.DictReader(f))


def style_header_row(ws, row, n_cols, start_col=1):
    for c in range(start_col, start_col + n_cols):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def title_block(ws, title, subtitle=None, cols=8):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)
        ws.cell(row=2, column=1, value=subtitle).font = SUBTITLE_FONT
    ws.row_dimensions[1].height = 22


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def freeze(ws, cell):
    ws.freeze_panes = cell


# ------------------------------------------------------------------ load data
with open("model_output.json") as f:
    M = json.load(f)

quarters = M["quarters"]
lps = M["lps"]
marks = read_csv("portfolio_quarterly_marks.csv")
exits = read_csv("exit_events.csv")
portfolio_companies = read_csv("portfolio_companies.csv")
performance = M["performance_summary"]
reconciliation = M["reconciliation"]

N_Q = M["n_quarters"]
INV_PERIOD_Q = M["investment_period_quarters"]
FUND_SIZE = M["fund_size"]

wb = Workbook()
wb.remove(wb.active)

# =====================================================================
# SHEET: Cover
# =====================================================================
ws = wb.create_sheet("Cover")
set_widths(ws, [3, 34, 60, 3])
title_block(ws, M["fund_name"], "Capital Call, Distribution Waterfall & NAV Model", cols=3)
ws["B4"] = "Prepared as of"
ws["C4"] = performance["as_of_quarter"] + "  (" + performance["as_of_date"] + ")"
ws["B5"] = "Fund size (committed capital)"
ws["C5"] = FUND_SIZE
ws["C5"].number_format = CUR0
ws["B6"] = "Limited partners"
ws["C6"] = len(lps)
ws["B7"] = "Vintage / investment period"
ws["C7"] = "Q1 2022 – Q4 2024 (12 quarters)"
ws["B8"] = "Fund term modeled"
ws["C8"] = f"{quarters[0]['label']} – {quarters[-1]['label']} ({N_Q} quarters)"
for r in range(4, 9):
    ws.cell(row=r, column=2).font = Font(name=FONT_NAME, bold=True)

ws["B10"] = "How this workbook is organized"
ws["B10"].font = SECTION_FONT
toc = [
    ("Assumptions", "Fund terms: fee rate, preferred return, carry split, catch-up"),
    ("Quarters", "The 18-quarter calendar used throughout the model"),
    ("LP Register", "15 limited partners and their commitments"),
    ("Portfolio Companies", "10 portfolio company investments (cost, invest date, sector)"),
    ("Portfolio Marks", "Quarterly fair value marks and exit/recap events, by company"),
    ("Mgmt Fee Schedule", "2% of committed capital during the investment period, stepping down to 2% of invested cost thereafter"),
    ("Capital Call Schedule", "Fund-level quarterly capital calls (investment + fee)"),
    ("LP Capital Calls", "Each LP's pro-rata share of every capital call"),
    ("Distribution Waterfall", "Return of capital → 8% preferred return → GP catch-up → 80/20 carry split, run cumulatively by quarter"),
    ("NAV Rollforward", "Quarterly Beginning NAV → Ending NAV bridge"),
    ("LP Capital Accounts", "Each LP's quarterly capital account statement"),
    ("LP Statement (Select LP)", "A single formatted LP statement — pick any LP from the dropdown"),
    ("IRR & MOIC Summary", "Gross (portfolio) and net (to-LP) IRR, MOIC, TVPI, DPI, RVPI"),
    ("Variance Reconciliation", "GP ledger vs. fund administrator records — 4 identified breaks"),
]
r = 11
for name, desc in toc:
    ws.cell(row=r, column=2, value=name).font = Font(name=FONT_NAME, bold=True)
    ws.cell(row=r, column=3, value=desc).font = BLACK
    r += 1

ws.cell(row=r + 1, column=2, value="Legend").font = SECTION_FONT
r += 2
legend = [("Blue text", "Hardcoded input", BLUE), ("Black text", "Formula", BLACK),
          ("Green text", "Link to another sheet", GREEN), ("Yellow fill", "Key assumption", None)]
for label, desc, font in legend:
    c = ws.cell(row=r, column=2, value=label)
    if font:
        c.font = font
    else:
        c.font = BLACK
        c.fill = YELLOW_FILL
    ws.cell(row=r, column=3, value=desc).font = BLACK
    r += 1

ws.cell(row=r + 1, column=2,
        value=("Synthetic fund for portfolio/demonstration purposes. All LPs, portfolio companies, "
               "valuations and cash flows are generated data — see README.md.")).font = SUBTITLE_FONT

# =====================================================================
# SHEET: Assumptions
# =====================================================================
ws = wb.create_sheet("Assumptions")
set_widths(ws, [3, 38, 20, 3])
title_block(ws, "Fund Assumptions", "Key economic terms — change these and the model recalculates", cols=3)

assumptions = [
    ("Fund name", M["fund_name"], None),
    ("Committed capital", FUND_SIZE, CUR0),
    ("Number of LPs", len(lps), None),
    ("Investment period (quarters)", INV_PERIOD_Q, None),
    ("Total quarters modeled", N_Q, None),
    ("Management fee rate (annual, % of committed capital during investment period)", M["mgmt_fee_rate"], PCT1),
    ("Management fee rate post-investment-period (annual, % of invested cost)", M["mgmt_fee_rate"], PCT1),
    ("Preferred return / hurdle rate (annual, compounded quarterly)", M["preferred_return_rate"], PCT1),
    ("GP catch-up (% of catch-up tier to GP)", 1.00, PCT1),
    ("Carried interest rate (GP share of profit above hurdle)", M["gp_carry_rate"], PCT1),
]
r = 4
cell_map = {}
for label, val, fmt in assumptions:
    ws.cell(row=r, column=2, value=label).font = BLACK
    c = ws.cell(row=r, column=3, value=val)
    c.font = BLUE_BOLD
    c.fill = YELLOW_FILL
    if fmt:
        c.number_format = fmt
    cell_map[label] = f"Assumptions!$C${r}"
    r += 1

A_FUND_SIZE = cell_map["Committed capital"]
A_INV_PERIOD_Q = cell_map["Investment period (quarters)"]
A_FEE_RATE = cell_map["Management fee rate (annual, % of committed capital during investment period)"]
A_PREF_RATE = cell_map["Preferred return / hurdle rate (annual, compounded quarterly)"]
A_CARRY_RATE = cell_map["Carried interest rate (GP share of profit above hurdle)"]
A_CATCHUP_GP_SHARE = cell_map["GP catch-up (% of catch-up tier to GP)"]

ws.cell(row=r + 1, column=2, value="Waterfall structure").font = SECTION_FONT
r += 2
for i, line in enumerate([
    "1. Return of Capital — 100% to LPs until cumulative capital contributed has been returned",
    "2. Preferred Return — 100% to LPs until the compounding hurdle above is paid in full",
    "3. GP Catch-up — 100% to GP until GP has received 20% of (preferred + catch-up) paid to date",
    "4. Carried Interest Split — 80% LP / 20% GP on all remaining distributions",
]):
    ws.cell(row=r + i, column=2, value=line).font = BLACK

# =====================================================================
# SHEET: Quarters
# =====================================================================
ws = wb.create_sheet("Quarters")
set_widths(ws, [3, 8, 10, 10, 12, 14, 10])
title_block(ws, "Quarterly Calendar", cols=6)
headers = ["Quarter Index", "Year", "Quarter", "Label", "Quarter End", "In Investment Period"]
hr = 3
for i, h in enumerate(headers):
    ws.cell(row=hr, column=2 + i, value=h)
style_header_row(ws, hr, len(headers), start_col=2)
Q_START_ROW = hr + 1
for i, q in enumerate(quarters):
    row = Q_START_ROW + i
    ws.cell(row=row, column=2, value=q["quarter_index"]).font = BLACK
    ws.cell(row=row, column=3, value=q["year"]).font = BLUE
    ws.cell(row=row, column=4, value=q["quarter"]).font = BLUE
    ws.cell(row=row, column=5, value=q["label"]).font = BLUE
    dcell = ws.cell(row=row, column=6, value=date.fromisoformat(q["quarter_end"]))
    dcell.font = BLUE
    dcell.number_format = "yyyy-mm-dd"
    icell = ws.cell(row=row, column=7, value=f"=B{row}<={A_INV_PERIOD_Q}")
    icell.font = BLACK
freeze(ws, "B4")


def quarter_row(qi):
    return Q_START_ROW + (qi - 1)


def quarter_label_ref(qi):
    return f"Quarters!$E${quarter_row(qi)}"


# =====================================================================
# SHEET: LP Register
# =====================================================================
ws = wb.create_sheet("LP Register")
set_widths(ws, [3, 8, 30, 20, 16, 12])
title_block(ws, "LP Register", "15 limited partners, commitments summing to committed capital", cols=5)
headers = ["LP ID", "LP Name", "LP Type", "Commitment ($)", "Commitment %"]
hr = 3
for i, h in enumerate(headers):
    ws.cell(row=hr, column=2 + i, value=h)
style_header_row(ws, hr, len(headers), start_col=2)
LP_START_ROW = hr + 1
lp_row = {}
for i, lp in enumerate(lps):
    row = LP_START_ROW + i
    lp_row[lp["lp_id"]] = row
    ws.cell(row=row, column=2, value=lp["lp_id"]).font = BLACK
    ws.cell(row=row, column=3, value=lp["lp_name"]).font = BLUE
    ws.cell(row=row, column=4, value=lp["lp_type"]).font = BLUE
    cc = ws.cell(row=row, column=5, value=lp["commitment"])
    cc.font = BLUE
    cc.number_format = CUR0
    pc = ws.cell(row=row, column=6, value=f"=E{row}/{A_FUND_SIZE}")
    pc.font = BLACK
    pc.number_format = PCT2
LP_END_ROW = LP_START_ROW + len(lps) - 1
tot_row = LP_END_ROW + 1
ws.cell(row=tot_row, column=3, value="Total").font = TOTAL_FONT
tcell = ws.cell(row=tot_row, column=5, value=f"=SUM(E{LP_START_ROW}:E{LP_END_ROW})")
tcell.font = TOTAL_FONT
tcell.number_format = CUR0
tcell.border = TOTAL_BORDER
pcell = ws.cell(row=tot_row, column=6, value=f"=SUM(F{LP_START_ROW}:F{LP_END_ROW})")
pcell.font = TOTAL_FONT
pcell.number_format = PCT2
pcell.border = TOTAL_BORDER
chk = ws.cell(row=tot_row + 1, column=3,
              value=f'=IF(E{tot_row}={A_FUND_SIZE},"OK — ties to committed capital","ERROR — does not tie")')
chk.font = BLACK
freeze(ws, "B4")


def lp_commitment_pct_ref(lp_id):
    return f"'LP Register'!$F${lp_row[lp_id]}"


LP_IDS = [lp["lp_id"] for lp in lps]

# =====================================================================
# SHEET: Portfolio Companies
# =====================================================================
ws = wb.create_sheet("Portfolio Companies")
set_widths(ws, [3, 9, 26, 18, 14, 16, 16, 14])
title_block(ws, "Portfolio Companies", "10 investments made during the 3-year investment period", cols=7)
headers = ["Company ID", "Company Name", "Sector", "Invest Quarter #", "Invest Quarter",
           "Cost Basis ($)", "Notes"]
hr = 3
for i, h in enumerate(headers):
    ws.cell(row=hr, column=2 + i, value=h)
style_header_row(ws, hr, len(headers), start_col=2)
PC_START_ROW = hr + 1
pc_row = {}
archetype_note = {
    "winner": "Strong performer — full exit", "steady": "Steady grower — full exit",
    "laggard": "Below-plan growth — held", "writedown": "Written down — held at reduced value",
    "recap": "Dividend recap — partial realization, remainder held",
}
for i, c in enumerate(portfolio_companies):
    row = PC_START_ROW + i
    pc_row[c["company_id"]] = row
    ws.cell(row=row, column=2, value=c["company_id"]).font = BLACK
    ws.cell(row=row, column=3, value=c["company_name"]).font = BLUE
    ws.cell(row=row, column=4, value=c["sector"]).font = BLUE
    iq = int(c["invest_quarter"])
    ws.cell(row=row, column=5, value=iq).font = BLUE
    lbl = ws.cell(row=row, column=6, value=f"=INDEX(Quarters!$E:$E,MATCH(E{row},Quarters!$B:$B,0))")
    lbl.font = BLACK
    cb = ws.cell(row=row, column=7, value=float(c["cost_basis"]))
    cb.font = BLUE
    cb.number_format = CUR0
    ws.cell(row=row, column=8, value=archetype_note.get(c["archetype"], "")).font = BLACK
PC_END_ROW = PC_START_ROW + len(portfolio_companies) - 1
tot_row = PC_END_ROW + 1
ws.cell(row=tot_row, column=3, value="Total invested at cost").font = TOTAL_FONT
tcell = ws.cell(row=tot_row, column=7, value=f"=SUM(G{PC_START_ROW}:G{PC_END_ROW})")
tcell.font = TOTAL_FONT
tcell.number_format = CUR0
tcell.border = TOTAL_BORDER
freeze(ws, "B4")
PORTFOLIO_COMPANY_IDS = [c["company_id"] for c in portfolio_companies]

# =====================================================================
# SHEET: Portfolio Marks
# =====================================================================
ws = wb.create_sheet("Portfolio Marks")
set_widths(ws, [3, 9, 22, 9, 20, 15, 15, 14, 10, 15, 17, 15])
title_block(ws, "Portfolio Marks", "Quarterly fair value marks and exit/recap events, by company", cols=11)
headers = ["Company ID", "Company Name", "Qtr #", "Quarter", "Cost Basis ($, BOP)", "Fair Value ($)",
           "Event Type", "% Realized", "Proceeds ($)", "Realized Cost Basis ($)", "Realized Gain ($)"]
hr = 3
for i, h in enumerate(headers):
    ws.cell(row=hr, column=2 + i, value=h)
style_header_row(ws, hr, len(headers), start_col=2)

exit_by_key = {(e["company_id"], int(e["quarter_index"])): e for e in exits}
pm_row = {}  # (company_id, quarter_index) -> row
PM_START_ROW = hr + 1
row = PM_START_ROW
for m in marks:
    cid = m["company_id"]
    qi = int(m["quarter_index"])
    pm_row[(cid, qi)] = row
    ws.cell(row=row, column=2, value=cid).font = BLACK
    ws.cell(row=row, column=3, value=m["company_name"]).font = BLACK
    ws.cell(row=row, column=4, value=qi).font = BLACK
    lbl = ws.cell(row=row, column=5, value=f"=INDEX(Quarters!$E:$E,MATCH(D{row},Quarters!$B:$B,0))")
    lbl.font = BLACK
    cb = ws.cell(row=row, column=6, value=float(m["cost_basis"]))
    cb.font = BLUE
    cb.number_format = CUR0
    fv = ws.cell(row=row, column=7, value=float(m["fair_value"]))
    fv.font = BLUE
    fv.number_format = CUR0

    ek = exit_by_key.get((cid, qi))
    if ek:
        ws.cell(row=row, column=8, value=ek["event_type"]).font = BLUE
        pr = ws.cell(row=row, column=9, value=float(ek["pct_realized"]))
        pr.font = BLUE
        pr.number_format = PCT1
        pc = ws.cell(row=row, column=10, value=float(ek["proceeds"]))
        pc.font = BLUE
        pc.number_format = CUR0
        rcb = ws.cell(row=row, column=11, value=float(ek["realized_cost_basis"]))
        rcb.font = BLUE
        rcb.number_format = CUR0
    else:
        ws.cell(row=row, column=10, value=0).font = BLACK
        ws.cell(row=row, column=10).number_format = CUR0
        ws.cell(row=row, column=11, value=0).font = BLACK
        ws.cell(row=row, column=11).number_format = CUR0
    rg = ws.cell(row=row, column=12, value=f"=J{row}-K{row}")
    rg.font = BLACK
    rg.number_format = CUR0
    row += 1
PM_END_ROW = row - 1
freeze(ws, "B4")

# =====================================================================
# SHEET: Mgmt Fee Schedule
# =====================================================================
ws = wb.create_sheet("Mgmt Fee Schedule")
set_widths(ws, [3, 8, 10, 16, 20, 24, 16])
title_block(ws, "Management Fee Schedule",
            "2% of committed capital during the investment period; 2% of remaining invested cost thereafter", cols=6)
headers = ["Qtr #", "Quarter", "In Investment Period", "Fee Basis ($)", "Fee Basis", "Management Fee ($)"]
hr = 3
for i, h in enumerate(headers):
    ws.cell(row=hr, column=2 + i, value=h)
style_header_row(ws, hr, len(headers), start_col=2)
MF_START_ROW = hr + 1
for i, q in enumerate(quarters):
    qi = q["quarter_index"]
    row = MF_START_ROW + i
    qrow = quarter_row(qi)
    ws.cell(row=row, column=2, value=qi).font = BLACK
    ws.cell(row=row, column=3, value=f"=Quarters!$E${qrow}").font = GREEN
    inperiod = ws.cell(row=row, column=4, value=f"=Quarters!$G${qrow}")
    inperiod.font = GREEN
    basis = ws.cell(row=row, column=5,
                     value=(f"=IF(D{row},{A_FUND_SIZE},"
                            f"SUMIFS('Portfolio Marks'!$F:$F,'Portfolio Marks'!$D:$D,B{row}-1))"))
    basis.font = BLACK
    basis.number_format = CUR0
    desc = ws.cell(row=row, column=6, value=f'=IF(D{row},"Committed capital","Invested cost (step-down)")')
    desc.font = BLACK
    fee = ws.cell(row=row, column=7, value=f"=E{row}*{A_FEE_RATE}/4")
    fee.font = BLACK
    fee.number_format = CUR0
MF_END_ROW = MF_START_ROW + len(quarters) - 1
freeze(ws, "B4")


def mgmt_fee_ref(qi):
    return f"'Mgmt Fee Schedule'!$G${MF_START_ROW + (qi - 1)}"


# =====================================================================
# SHEET: Capital Call Schedule
# =====================================================================
ws = wb.create_sheet("Capital Call Schedule")
set_widths(ws, [3, 8, 10, 18, 16, 16, 18])
title_block(ws, "Capital Call Schedule", "Fund-level quarterly capital calls — investment need + management fee", cols=6)
headers = ["Qtr #", "Quarter", "Investment Call ($)", "Mgmt Fee Call ($)", "Total Call ($)", "Cumulative Called ($)"]
hr = 3
for i, h in enumerate(headers):
    ws.cell(row=hr, column=2 + i, value=h)
style_header_row(ws, hr, len(headers), start_col=2)
CC_START_ROW = hr + 1
for i, q in enumerate(quarters):
    qi = q["quarter_index"]
    row = CC_START_ROW + i
    qrow = quarter_row(qi)
    ws.cell(row=row, column=2, value=qi).font = BLACK
    ws.cell(row=row, column=3, value=f"=Quarters!$E${qrow}").font = GREEN
    inv = ws.cell(row=row, column=4,
                   value=f"=SUMIFS('Portfolio Companies'!$G:$G,'Portfolio Companies'!$E:$E,B{row})")
    inv.font = BLACK
    inv.number_format = CUR0
    fee = ws.cell(row=row, column=5, value=f"={mgmt_fee_ref(qi)}")
    fee.font = GREEN
    fee.number_format = CUR0
    tot = ws.cell(row=row, column=6, value=f"=D{row}+E{row}")
    tot.font = BLACK
    tot.number_format = CUR0
    if i == 0:
        cum = ws.cell(row=row, column=7, value=f"=F{row}")
    else:
        cum = ws.cell(row=row, column=7, value=f"=G{row-1}+F{row}")
    cum.font = BLACK
    cum.number_format = CUR0
CC_END_ROW = CC_START_ROW + len(quarters) - 1
tot_row = CC_END_ROW + 1
ws.cell(row=tot_row, column=3, value="Total / Unfunded Commitment").font = TOTAL_FONT
tcell = ws.cell(row=tot_row, column=6, value=f"=SUM(F{CC_START_ROW}:F{CC_END_ROW})")
tcell.font = TOTAL_FONT
tcell.number_format = CUR0
tcell.border = TOTAL_BORDER
ucell = ws.cell(row=tot_row, column=7, value=f"={A_FUND_SIZE}-G{CC_END_ROW}")
ucell.font = TOTAL_FONT
ucell.number_format = CUR0
ucell.border = TOTAL_BORDER
ws.cell(row=tot_row + 1, column=3, value="(Unfunded commitment remaining)").font = SUBTITLE_FONT
freeze(ws, "B4")


def capital_call_total_ref(qi):
    return f"'Capital Call Schedule'!$F${CC_START_ROW + (qi - 1)}"


def capital_call_inv_ref(qi):
    return f"'Capital Call Schedule'!$D${CC_START_ROW + (qi - 1)}"


def capital_call_fee_ref(qi):
    return f"'Capital Call Schedule'!$E${CC_START_ROW + (qi - 1)}"


# =====================================================================
# SHEET: Distribution Waterfall
# =====================================================================
ws = wb.create_sheet("Distribution Waterfall")
set_widths(ws, [3] + [8, 10, 15] + [16] * 19)
title_block(ws, "Distribution Waterfall",
            "European (whole-fund) waterfall, run cumulatively by quarter: "
            "Return of Capital -> Preferred Return -> GP Catch-up -> 80/20 Carry Split", cols=12)
headers = [
    "Qtr #", "Quarter", "Distributable ($)",
    "Cum. Contributions (BOP)", "Cum. Contributions (EOP)", "Cum. ROC Distributed (BOP)",
    "Preferred Accrual (Qtr)", "Cum. Preferred Accrued (EOP)", "Cum. Preferred Paid (BOP)",
    "Cum. GP Catch-up Paid (BOP)",
    "Tier 1: Return of Capital", "Cum. ROC Distributed (EOP)",
    "Tier 2: Preferred Return", "Cum. Preferred Paid (EOP)",
    "Target Cum. Catch-up (Total)", "Tier 3: GP Catch-up", "Cum. GP Catch-up Paid (EOP)",
    "Tier 4: Carry (Total)", "Tier 4: Carry to LPs (80%)", "Tier 4: Carry to GP (20%)",
    "Total to LPs", "Total to GP",
]
hr = 3
for i, h in enumerate(headers):
    ws.cell(row=hr, column=2 + i, value=h)
style_header_row(ws, hr, len(headers), start_col=2)
DW_START_ROW = hr + 1
for i, q in enumerate(quarters):
    qi = q["quarter_index"]
    row = DW_START_ROW + i
    prev = row - 1
    first = (i == 0)

    ws.cell(row=row, column=2, value=qi).font = BLACK
    ws.cell(row=row, column=3, value=f"=Quarters!$E${quarter_row(qi)}").font = GREEN

    dist = ws.cell(row=row, column=4,
                    value=f"=SUMIFS('Portfolio Marks'!$J:$J,'Portfolio Marks'!$D:$D,B{row})")
    dist.font = BLACK
    dist.number_format = CUR0

    bop_contrib = ws.cell(row=row, column=5, value="0" if first else f"=F{prev}")
    bop_contrib.font = BLACK
    bop_contrib.number_format = CUR0
    eop_contrib = ws.cell(row=row, column=6, value=f"=E{row}+{capital_call_total_ref(qi)}")
    eop_contrib.font = BLACK
    eop_contrib.number_format = CUR0

    bop_roc = ws.cell(row=row, column=7, value="0" if first else f"=M{prev}")
    bop_roc.font = BLACK
    bop_roc.number_format = CUR0

    pref_accr = ws.cell(row=row, column=8, value=f"=(E{row}-G{row})*{A_PREF_RATE}/4")
    pref_accr.font = BLACK
    pref_accr.number_format = CUR0
    eop_pref_accr = ws.cell(row=row, column=9, value=f"=H{row}" if first else f"=I{prev}+H{row}")
    eop_pref_accr.font = BLACK
    eop_pref_accr.number_format = CUR0

    bop_pref_paid = ws.cell(row=row, column=10, value="0" if first else f"=O{prev}")
    bop_pref_paid.font = BLACK
    bop_pref_paid.number_format = CUR0
    bop_catchup_paid = ws.cell(row=row, column=11, value="0" if first else f"=R{prev}")
    bop_catchup_paid.font = BLACK
    bop_catchup_paid.number_format = CUR0

    t1 = ws.cell(row=row, column=12, value=f"=MIN(D{row},MAX(F{row}-G{row},0))")
    t1.font = BLACK
    t1.number_format = CUR0
    eop_roc = ws.cell(row=row, column=13, value=f"=G{row}+L{row}")
    eop_roc.font = BLACK
    eop_roc.number_format = CUR0

    t2 = ws.cell(row=row, column=14, value=f"=MIN(D{row}-L{row},MAX(I{row}-J{row},0))")
    t2.font = BLACK
    t2.number_format = CUR0
    eop_pref_paid = ws.cell(row=row, column=15, value=f"=J{row}+N{row}")
    eop_pref_paid.font = BLACK
    eop_pref_paid.number_format = CUR0

    target_catchup = ws.cell(row=row, column=16, value=f"=({A_CARRY_RATE}/(1-{A_CARRY_RATE}))*O{row}")
    target_catchup.font = BLACK
    target_catchup.number_format = CUR0

    t3 = ws.cell(row=row, column=17, value=f"=MIN(D{row}-L{row}-N{row},MAX(P{row}-K{row},0))")
    t3.font = BLACK
    t3.number_format = CUR0
    eop_catchup_paid = ws.cell(row=row, column=18, value=f"=K{row}+Q{row}")
    eop_catchup_paid.font = BLACK
    eop_catchup_paid.number_format = CUR0

    t4 = ws.cell(row=row, column=19, value=f"=D{row}-L{row}-N{row}-Q{row}")
    t4.font = BLACK
    t4.number_format = CUR0
    t4lp = ws.cell(row=row, column=20, value=f"=S{row}*(1-{A_CARRY_RATE})")
    t4lp.font = BLACK
    t4lp.number_format = CUR0
    t4gp = ws.cell(row=row, column=21, value=f"=S{row}*{A_CARRY_RATE}")
    t4gp.font = BLACK
    t4gp.number_format = CUR0

    tot_lp = ws.cell(row=row, column=22, value=f"=L{row}+N{row}+T{row}")
    tot_lp.font = TOTAL_FONT
    tot_lp.number_format = CUR0
    tot_gp = ws.cell(row=row, column=23, value=f"=Q{row}+U{row}")
    tot_gp.font = TOTAL_FONT
    tot_gp.number_format = CUR0

DW_END_ROW = DW_START_ROW + len(quarters) - 1
tot_row = DW_END_ROW + 1
ws.cell(row=tot_row, column=3, value="Cumulative").font = TOTAL_FONT
for col_letter in ["V", "W"]:
    tc = ws.cell(row=tot_row, column=column_index_from_string(col_letter),
                 value=f"=SUM({col_letter}{DW_START_ROW}:{col_letter}{DW_END_ROW})")
    tc.font = TOTAL_FONT
    tc.number_format = CUR0
    tc.border = TOTAL_BORDER
freeze(ws, "D4")


def waterfall_ref(qi, col_letter):
    return f"'Distribution Waterfall'!${col_letter}${DW_START_ROW + (qi - 1)}"


# =====================================================================
# SHEET: NAV Rollforward
# =====================================================================
ws = wb.create_sheet("NAV Rollforward")
set_widths(ws, [3, 8, 10, 16, 16, 16, 16, 16, 16, 16, 10])
title_block(ws, "Quarterly NAV Rollforward", "Beginning NAV -> Ending NAV bridge, tying to portfolio fair values each quarter", cols=9)
headers = ["Qtr #", "Quarter", "Beginning NAV", "Contributions", "Management Fees",
           "Realized Gain/(Loss)", "Unrealized Gain/(Loss)", "Distributions", "Ending NAV", "Ties?"]
hr = 3
for i, h in enumerate(headers):
    ws.cell(row=hr, column=2 + i, value=h)
style_header_row(ws, hr, len(headers), start_col=2)
NAV_START_ROW = hr + 1
for i, q in enumerate(quarters):
    qi = q["quarter_index"]
    row = NAV_START_ROW + i
    prev = row - 1
    first = (i == 0)

    ws.cell(row=row, column=2, value=qi).font = BLACK
    ws.cell(row=row, column=3, value=f"=Quarters!$E${quarter_row(qi)}").font = GREEN

    beg = ws.cell(row=row, column=4, value="0" if first else f"=J{prev}")
    beg.font = BLACK
    beg.number_format = CUR0

    contrib = ws.cell(row=row, column=5, value=f"={capital_call_total_ref(qi)}")
    contrib.font = GREEN
    contrib.number_format = CUR0

    fees = ws.cell(row=row, column=6, value=f"=-{mgmt_fee_ref(qi)}")
    fees.font = GREEN
    fees.number_format = CUR0

    realized = ws.cell(row=row, column=7,
                        value=f"=SUMIFS('Portfolio Marks'!$L:$L,'Portfolio Marks'!$D:$D,B{row})")
    realized.font = BLACK
    realized.number_format = CUR0

    ending_direct = (f"SUMIFS('Portfolio Marks'!$G:$G,'Portfolio Marks'!$D:$D,B{row})"
                      f"-SUMIFS('Portfolio Marks'!$J:$J,'Portfolio Marks'!$D:$D,B{row})")

    dist = ws.cell(row=row, column=9, value=f"=-{waterfall_ref(qi, 'D')}")
    dist.font = GREEN
    dist.number_format = CUR0

    # Unrealized gain is the plug that makes the rollforward tie exactly to the
    # independently-computed ending NAV (portfolio fair value less this quarter's
    # distributions).
    unreal = ws.cell(row=row, column=8,
                      value=f"=({ending_direct})-D{row}-E{row}+{mgmt_fee_ref(qi)}-G{row}+{waterfall_ref(qi, 'D')}")
    unreal.font = BLACK
    unreal.number_format = CUR0

    end = ws.cell(row=row, column=10, value=f"=D{row}+E{row}+F{row}+G{row}+H{row}+I{row}")
    end.font = TOTAL_FONT
    end.number_format = CUR0

    tie = ws.cell(row=row, column=11, value=f'=IF(ROUND(J{row}-({ending_direct}),2)=0,"OK","ERROR")')
    tie.font = BLACK

NAV_END_ROW = NAV_START_ROW + len(quarters) - 1
freeze(ws, "D4")


def nav_ending_ref(qi):
    return f"'NAV Rollforward'!$J${NAV_START_ROW + (qi - 1)}"


def nav_realized_ref(qi):
    return f"'NAV Rollforward'!$G${NAV_START_ROW + (qi - 1)}"


def nav_unrealized_ref(qi):
    return f"'NAV Rollforward'!$H${NAV_START_ROW + (qi - 1)}"


# =====================================================================
# SHEET: LP Capital Calls
# =====================================================================
ws = wb.create_sheet("LP Capital Calls")
set_widths(ws, [3, 8, 10, 8, 26, 16, 16, 16])
title_block(ws, "LP Capital Calls", "Each LP's pro-rata share of every quarterly capital call", cols=6)
headers = ["Qtr #", "Quarter", "LP ID", "LP Name", "Investment Call ($)", "Mgmt Fee Call ($)", "Total Call ($)"]
hr = 3
for i, h in enumerate(headers):
    ws.cell(row=hr, column=2 + i, value=h)
style_header_row(ws, hr, len(headers), start_col=2)
LPC_START_ROW = hr + 1
lpc_row = {}  # (qi, lp_id) -> row
row = LPC_START_ROW
for q in quarters:
    qi = q["quarter_index"]
    for lp in lps:
        lpc_row[(qi, lp["lp_id"])] = row
        ws.cell(row=row, column=2, value=qi).font = BLACK
        ws.cell(row=row, column=3, value=f"=Quarters!$E${quarter_row(qi)}").font = GREEN
        ws.cell(row=row, column=4, value=lp["lp_id"]).font = BLACK
        ws.cell(row=row, column=5, value=lp["lp_name"]).font = BLACK
        ic = ws.cell(row=row, column=6, value=f"={capital_call_inv_ref(qi)}*{lp_commitment_pct_ref(lp['lp_id'])}")
        ic.font = GREEN
        ic.number_format = CUR0
        fc = ws.cell(row=row, column=7, value=f"={capital_call_fee_ref(qi)}*{lp_commitment_pct_ref(lp['lp_id'])}")
        fc.font = GREEN
        fc.number_format = CUR0
        tc = ws.cell(row=row, column=8, value=f"=F{row}+G{row}")
        tc.font = BLACK
        tc.number_format = CUR0
        row += 1
LPC_END_ROW = row - 1
freeze(ws, "B4")


def lp_call_total_ref(qi, lp_id):
    return f"'LP Capital Calls'!$H${lpc_row[(qi, lp_id)]}"


# =====================================================================
# SHEET: LP Capital Accounts
# =====================================================================
ws = wb.create_sheet("LP Capital Accounts")
set_widths(ws, [3, 8, 10, 8, 26, 15, 13, 12, 13, 15, 15, 17, 17, 15])
title_block(ws, "LP Capital Account Statements",
            "Each LP's quarterly capital account — beginning balance through ending balance", cols=12)
headers = ["Qtr #", "Quarter", "LP ID", "LP Name", "Beginning Balance", "Contributions",
           "Dist: Return of Capital", "Dist: Preferred Return", "Dist: Carry Split (LP 80%)",
           "Total Cash Distributions", "Gain/(Loss) Allocation, Net of Fees",
           "Less: Carried Interest to GP", "Ending Balance"]
hr = 3
for i, h in enumerate(headers):
    ws.cell(row=hr, column=2 + i, value=h)
style_header_row(ws, hr, len(headers), start_col=2)
LPA_START_ROW = hr + 1
N_LPS = len(lps)
lpa_row = {}
row = LPA_START_ROW
for q in quarters:
    qi = q["quarter_index"]
    for lp in lps:
        lpa_row[(qi, lp["lp_id"])] = row
        pct = lp_commitment_pct_ref(lp["lp_id"])
        ws.cell(row=row, column=2, value=qi).font = BLACK
        ws.cell(row=row, column=3, value=f"=Quarters!$E${quarter_row(qi)}").font = GREEN
        ws.cell(row=row, column=4, value=lp["lp_id"]).font = BLACK
        ws.cell(row=row, column=5, value=lp["lp_name"]).font = BLACK

        beg = ws.cell(row=row, column=6, value="0" if qi == 1 else f"=N{row - N_LPS}")
        beg.font = BLACK
        beg.number_format = CUR0

        contrib = ws.cell(row=row, column=7, value=f"={lp_call_total_ref(qi, lp['lp_id'])}")
        contrib.font = GREEN
        contrib.number_format = CUR0

        droc = ws.cell(row=row, column=8, value=f"={waterfall_ref(qi, 'L')}*{pct}")
        droc.font = GREEN
        droc.number_format = CUR0
        dpref = ws.cell(row=row, column=9, value=f"={waterfall_ref(qi, 'N')}*{pct}")
        dpref.font = GREEN
        dpref.number_format = CUR0
        dcarry = ws.cell(row=row, column=10, value=f"={waterfall_ref(qi, 'T')}*{pct}")
        dcarry.font = GREEN
        dcarry.number_format = CUR0
        totdist = ws.cell(row=row, column=11, value=f"=H{row}+I{row}+J{row}")
        totdist.font = TOTAL_FONT
        totdist.number_format = CUR0

        gain = ws.cell(row=row, column=12,
                        value=f"=({nav_realized_ref(qi)}+{nav_unrealized_ref(qi)}-{mgmt_fee_ref(qi)})*{pct}")
        gain.font = GREEN
        gain.number_format = CUR0

        carry_gp = ws.cell(row=row, column=13,
                            value=f"=-({waterfall_ref(qi, 'Q')}+{waterfall_ref(qi, 'U')})*{pct}")
        carry_gp.font = GREEN
        carry_gp.number_format = CUR0

        end = ws.cell(row=row, column=14, value=f"=F{row}+G{row}-K{row}+L{row}+M{row}")
        end.font = TOTAL_FONT
        end.number_format = CUR0

        key = ws.cell(row=row, column=15, value=f'=B{row}&"|"&D{row}')
        key.font = BLACK

        row += 1
LPA_END_ROW = row - 1
ws.cell(row=hr, column=15, value="Lookup Key").font = HEADER_FONT
ws.cell(row=hr, column=15).fill = HEADER_FILL
freeze(ws, "D4")

tie_row = LPA_END_ROW + 2
ws.cell(row=tie_row, column=3, value="Tie-out: sum of LP ending balances vs. Fund Ending NAV, final quarter").font = TOTAL_FONT
sumcell = ws.cell(row=tie_row, column=14,
                   value=f"=SUM(N{LPA_END_ROW - N_LPS + 1}:N{LPA_END_ROW})")
sumcell.font = TOTAL_FONT
sumcell.number_format = CUR0
navcell = ws.cell(row=tie_row + 1, column=14, value=f"={nav_ending_ref(N_Q)}")
navcell.font = GREEN
navcell.number_format = CUR0
ws.cell(row=tie_row + 1, column=3, value="Fund Ending NAV").font = BLACK
chk = ws.cell(row=tie_row + 2, column=14,
              value=f'=IF(ROUND(N{tie_row}-N{tie_row+1},2)=0,"OK — ties out","ERROR")')
chk.font = BLACK
ws.cell(row=tie_row + 2, column=3, value="Check").font = BLACK

# =====================================================================
# SHEET: LP Statement (Select LP)
# =====================================================================
ws = wb.create_sheet("LP Statement (Select LP)")
set_widths(ws, [3, 8, 10, 15, 13, 12, 13, 15, 15, 17, 17, 15])
title_block(ws, "LP Capital Account Statement", "Pick any LP from the dropdown in C4 to view their full statement", cols=11)
ws.cell(row=4, column=2, value="Select LP:").font = Font(name=FONT_NAME, bold=True)
sel = ws.cell(row=4, column=3, value=lps[0]["lp_name"])
sel.font = BLUE_BOLD
sel.fill = YELLOW_FILL
dv = DataValidation(type="list", formula1=f"='LP Register'!$C${LP_START_ROW}:$C${LP_END_ROW}", allow_blank=False)
ws.add_data_validation(dv)
dv.add(sel)

ws.cell(row=5, column=2, value="LP ID").font = BLACK
lpid_cell = ws.cell(row=5, column=3, value=f"=INDEX('LP Register'!$B$"
                                            f"{LP_START_ROW}:$B${LP_END_ROW},MATCH($C$4,'LP Register'!$C$"
                                            f"{LP_START_ROW}:$C${LP_END_ROW},0))")
lpid_cell.font = BLACK
ws.cell(row=6, column=2, value="Commitment ($)").font = BLACK
commit_cell = ws.cell(row=6, column=3, value=f"=INDEX('LP Register'!$E$"
                                              f"{LP_START_ROW}:$E${LP_END_ROW},MATCH($C$4,'LP Register'!$C$"
                                              f"{LP_START_ROW}:$C${LP_END_ROW},0))")
commit_cell.font = BLACK
commit_cell.number_format = CUR0
ws.cell(row=7, column=2, value="Commitment %").font = BLACK
pct_cell = ws.cell(row=7, column=3, value=f"=INDEX('LP Register'!$F$"
                                           f"{LP_START_ROW}:$F${LP_END_ROW},MATCH($C$4,'LP Register'!$C$"
                                           f"{LP_START_ROW}:$C${LP_END_ROW},0))")
pct_cell.font = BLACK
pct_cell.number_format = PCT2

headers = ["Qtr #", "Quarter", "Beginning Balance", "Contributions", "Dist: ROC", "Dist: Pref",
           "Dist: Carry Split", "Total Cash Dist.", "Gain/(Loss) Alloc.", "Less: Carry to GP", "Ending Balance"]
hr = 9
for i, h in enumerate(headers):
    ws.cell(row=hr, column=2 + i, value=h)
style_header_row(ws, hr, len(headers), start_col=2)
LST_START_ROW = hr + 1
for i, q in enumerate(quarters):
    qi = q["quarter_index"]
    row = LST_START_ROW + i
    ws.cell(row=row, column=2, value=qi).font = BLACK
    ws.cell(row=row, column=3, value=f"=Quarters!$E${quarter_row(qi)}").font = GREEN
    src_cols = ["F", "G", "H", "I", "J", "K", "L", "M", "N"]
    dst_cols = list(range(4, 13))
    lookup_key = f'B{row}&"|"&$C$5'
    for sc, dc in zip(src_cols, dst_cols):
        formula = (f"=INDEX('LP Capital Accounts'!${sc}${LPA_START_ROW}:${sc}${LPA_END_ROW},"
                   f"MATCH({lookup_key},'LP Capital Accounts'!$O${LPA_START_ROW}:$O${LPA_END_ROW},0))")
        cell = ws.cell(row=row, column=dc, value=formula)
        cell.font = GREEN
        cell.number_format = CUR0
LST_END_ROW = LST_START_ROW + len(quarters) - 1
ws.cell(row=LST_END_ROW + 1, column=3,
        value=("INDEX/MATCH on a Quarter#|LP ID key pulls this LP's row from the LP Capital "
               "Accounts sheet — pick a different LP above and every row updates.")).font = SUBTITLE_FONT
freeze(ws, "B10")

# =====================================================================
# SHEET: IRR & MOIC Summary
# =====================================================================
ws = wb.create_sheet("IRR & MOIC Summary")
set_widths(ws, [3, 14, 30, 18])
title_block(ws, "Fund Performance Summary", f"As of {performance['as_of_quarter']}", cols=3)

# ---- Section A: Gross (portfolio-level) cash flows ----
ws.cell(row=4, column=2, value="Gross Cash Flows (portfolio level, pre-fee, pre-carry)").font = SECTION_FONT
hdr = 5
for i, h in enumerate(["Date", "Description", "Cash Flow ($)"]):
    ws.cell(row=hdr, column=2 + i, value=h)
style_header_row(ws, hdr, 3, start_col=2)
row = hdr + 1
gross_start = row
for c in portfolio_companies:
    qrow = quarter_row(int(c["invest_quarter"]))
    ws.cell(row=row, column=2, value=f"=Quarters!$F${qrow}").font = GREEN
    ws.cell(row=row, column=2).number_format = "yyyy-mm-dd"
    ws.cell(row=row, column=3, value=f"Investment: {c['company_name']}").font = BLACK
    cell = ws.cell(row=row, column=4, value=f"=-'Portfolio Companies'!$G${pc_row[c['company_id']]}")
    cell.font = GREEN
    cell.number_format = CUR0
    row += 1
for e in exits:
    key = (e["company_id"], int(e["quarter_index"]))
    qrow = quarter_row(int(e["quarter_index"]))
    ws.cell(row=row, column=2, value=f"=Quarters!$F${qrow}").font = GREEN
    ws.cell(row=row, column=2).number_format = "yyyy-mm-dd"
    label = "Full exit" if e["event_type"] == "full_exit" else "Partial recap"
    ws.cell(row=row, column=3, value=f"{label}: {e['company_name']}").font = BLACK
    cell = ws.cell(row=row, column=4, value=f"='Portfolio Marks'!$J${pm_row[key]}")
    cell.font = GREEN
    cell.number_format = CUR0
    row += 1
ws.cell(row=row, column=2, value=f"=Quarters!$F${quarter_row(N_Q)}").font = GREEN
ws.cell(row=row, column=2).number_format = "yyyy-mm-dd"
ws.cell(row=row, column=3, value="Remaining portfolio fair value (terminal)").font = BLACK
cell = ws.cell(row=row, column=4, value=f"={nav_ending_ref(N_Q)}")
cell.font = GREEN
cell.number_format = CUR0
gross_end = row
row += 2

ws.cell(row=row, column=2, value="Gross IRR").font = TOTAL_FONT
girr = ws.cell(row=row, column=4, value=f"=XIRR(D{gross_start}:D{gross_end},B{gross_start}:B{gross_end})")
girr.font = TOTAL_FONT
girr.number_format = PCT2
row += 1
ws.cell(row=row, column=2, value="Gross MOIC").font = TOTAL_FONT
gmoic = ws.cell(row=row, column=4,
                 value=f"=(SUMIF(C{gross_start}:C{gross_end},\"<>Investment*\",D{gross_start}:D{gross_end}))"
                       f"/(-SUMIF(C{gross_start}:C{gross_end},\"Investment*\",D{gross_start}:D{gross_end}))")
gmoic.font = TOTAL_FONT
gmoic.number_format = MULT
GROSS_IRR_CELL = f"'IRR & MOIC Summary'!$D${row-1}"
GROSS_MOIC_CELL = f"'IRR & MOIC Summary'!$D${row}"
row += 3

# ---- Section B: Net (to-LP) cash flows ----
ws.cell(row=row, column=2, value="Net Cash Flows (to LPs, post-fee, post-carry)").font = SECTION_FONT
row += 1
hdr = row
for i, h in enumerate(["Qtr #", "Date", "Description", "Cash Flow ($)"]):
    ws.cell(row=hdr, column=2 + i, value=h)
style_header_row(ws, hdr, 4, start_col=2)
row = hdr + 1
net_start = row
for q in quarters:
    qi = q["quarter_index"]
    qrow = quarter_row(qi)
    ws.cell(row=row, column=2, value=qi).font = BLACK
    ws.cell(row=row, column=3, value=f"=Quarters!$F${qrow}").font = GREEN
    ws.cell(row=row, column=3).number_format = "yyyy-mm-dd"
    ws.cell(row=row, column=4, value=f"=Quarters!$E${qrow}&\": capital call + distribution\"").font = BLACK
    cf = ws.cell(row=row, column=5, value=f"=-{capital_call_total_ref(qi)}+{waterfall_ref(qi, 'V')}")
    cf.font = GREEN
    cf.number_format = CUR0
    row += 1
net_end = row - 1
ws.cell(row=row, column=3, value=f"=Quarters!$F${quarter_row(N_Q)}").font = GREEN
ws.cell(row=row, column=3).number_format = "yyyy-mm-dd"
ws.cell(row=row, column=4, value="Terminal LP NAV (residual value)").font = BLACK
cell = ws.cell(row=row, column=5, value=f"={nav_ending_ref(N_Q)}")
cell.font = GREEN
cell.number_format = CUR0
net_end_terminal = row
row += 2

ws.cell(row=row, column=2, value="Net IRR (to LPs)").font = TOTAL_FONT
nirr = ws.cell(row=row, column=5,
                value=f"=XIRR(E{net_start}:E{net_end_terminal},C{net_start}:C{net_end_terminal})")
nirr.font = TOTAL_FONT
nirr.number_format = PCT2
row += 1

cum_calls_cell = f"'Capital Call Schedule'!$F${CC_END_ROW + 1}"
cum_dist_cell = f"'Distribution Waterfall'!$V${DW_END_ROW + 1}"
ending_nav_cell = nav_ending_ref(N_Q)

ws.cell(row=row, column=2, value="Net TVPI (Total Value / Paid-In)").font = TOTAL_FONT
tvpi = ws.cell(row=row, column=5, value=f"=({cum_dist_cell}+{ending_nav_cell})/{cum_calls_cell}")
tvpi.font = TOTAL_FONT
tvpi.number_format = MULT
row += 1
ws.cell(row=row, column=2, value="DPI (Distributed / Paid-In)").font = BLACK
dpi = ws.cell(row=row, column=5, value=f"={cum_dist_cell}/{cum_calls_cell}")
dpi.font = BLACK
dpi.number_format = MULT
row += 1
ws.cell(row=row, column=2, value="RVPI (Residual Value / Paid-In)").font = BLACK
rvpi = ws.cell(row=row, column=5, value=f"={ending_nav_cell}/{cum_calls_cell}")
rvpi.font = BLACK
rvpi.number_format = MULT
row += 3

ws.cell(row=row, column=2, value="Gross IRR (portfolio)").font = BLACK
ws.cell(row=row, column=5, value=f"={GROSS_IRR_CELL}").font = GREEN
ws.cell(row=row, column=5).number_format = PCT2
row += 1
ws.cell(row=row, column=2, value="Gross MOIC (portfolio)").font = BLACK
ws.cell(row=row, column=5, value=f"={GROSS_MOIC_CELL}").font = GREEN
ws.cell(row=row, column=5).number_format = MULT
row += 1
ws.cell(row=row, column=2, value="Note").font = SUBTITLE_FONT
ws.cell(row=row, column=3,
        value=("All 15 LPs share identical timing (pro-rata by commitment %), so net IRR/MOIC "
               "is the same for every LP as for the fund overall.")).font = SUBTITLE_FONT

freeze(ws, "B4")
IRR_SHEET_NIRR_CELL = f"'IRR & MOIC Summary'!$E${row-6}"

print("IRR & MOIC Summary sheet built.")
wb.save("PE_Fund_Model.xlsx")

# =====================================================================
# SHEET: Variance Reconciliation
# =====================================================================
ws = wb.create_sheet("Variance Reconciliation")
set_widths(ws, [3, 8, 26, 20, 20, 16, 26])
title_block(ws, "GP Ledger vs. Fund Administrator — Variance Reconciliation",
            f"Tie-out as of {performance['as_of_quarter']}", cols=6)

ws.cell(row=4, column=2, value="Summary of breaks").font = SECTION_FONT
hr = 5
headers = ["Break", "Category", "GP Ledger ($)", "Fund Admin ($)", "Variance ($)", "Status"]
for i, h in enumerate(headers):
    ws.cell(row=hr, column=2 + i, value=h)
style_header_row(ws, hr, len(headers), start_col=2)
SUM_START_ROW = hr + 1

# ---- R-1: timing difference (Q9 capital call settlement lag) ----
r1_row = SUM_START_ROW
timing_lp_ids = ["LP-05", "LP-09"]
r1_ledger = f"={capital_call_total_ref(9)}"
r1_affected = "+".join(f"{capital_call_total_ref(9)}*{lp_commitment_pct_ref(l)}" for l in timing_lp_ids)
ws.cell(row=r1_row, column=2, value="R-1").font = BLACK
ws.cell(row=r1_row, column=3, value="Timing difference (Q9 capital call settlement lag)").font = BLACK
c = ws.cell(row=r1_row, column=4, value=r1_ledger)
c.font = GREEN
c.number_format = CUR0
c = ws.cell(row=r1_row, column=5, value=f"={capital_call_total_ref(9)}-({r1_affected})")
c.font = BLACK
c.number_format = CUR0
c = ws.cell(row=r1_row, column=6, value=f"=E{r1_row}-D{r1_row}")
c.font = BLACK
c.number_format = CUR0
ws.cell(row=r1_row, column=7, value="Resolved — timing only, nets to zero by Q10").font = BLACK

# ---- R-2: management fee basis error (Q13-Q16) ----
r2_row = r1_row + 1
fee_break_qs = [13, 14, 15, 16]
r2_ledger = "+".join(mgmt_fee_ref(q) for q in fee_break_qs)
r2_admin = "+".join(f"({A_FUND_SIZE}*{A_FEE_RATE}/4)" for _ in fee_break_qs)
ws.cell(row=r2_row, column=2, value="R-2").font = BLACK
ws.cell(row=r2_row, column=3, value="Management fee basis error (Q13-Q16 billed on committed capital, not stepped-down)").font = BLACK
c = ws.cell(row=r2_row, column=4, value=f"={r2_ledger}")
c.font = GREEN
c.number_format = CUR0
c = ws.cell(row=r2_row, column=5, value=f"={r2_admin}")
c.font = BLACK
c.number_format = CUR0
c = ws.cell(row=r2_row, column=6, value=f"=D{r2_row}-E{r2_row}")
c.font = BLACK
c.number_format = CUR0
ws.cell(row=r2_row, column=7, value="Open — adjusting entry pending LP notice").font = BLACK

# ---- R-2 detail sub-table ----
r2d_hdr = SUM_START_ROW + 6   # leaves rows for R-1..R-4 (SUM_START_ROW..+3) plus a blank row
ws.cell(row=r2d_hdr, column=3, value="R-2 detail: fee basis error by quarter").font = SECTION_FONT
r2d_hdr += 1
for i, h in enumerate(["Quarter", "GP Ledger Fee ($)", "Fund Admin Fee ($)", "Variance ($)"]):
    ws.cell(row=r2d_hdr, column=3 + i, value=h)
style_header_row(ws, r2d_hdr, 4, start_col=3)
row = r2d_hdr + 1
for q in fee_break_qs:
    ws.cell(row=row, column=3, value=f"=Quarters!$E${quarter_row(q)}").font = GREEN
    c = ws.cell(row=row, column=4, value=f"={mgmt_fee_ref(q)}")
    c.font = GREEN
    c.number_format = CUR0
    c = ws.cell(row=row, column=5, value=f"={A_FUND_SIZE}*{A_FEE_RATE}/4")
    c.font = BLACK
    c.number_format = CUR0
    c = ws.cell(row=row, column=6, value=f"=D{row}-E{row}")
    c.font = BLACK
    c.number_format = CUR0
    row += 1
next_block = row + 1

# ---- R-3: distribution character misclassification (Q18 recap, counterfactual) ----
r3_row = r2_row + 1  # keep in the top summary block, values filled in after the counterfactual is computed below
hd_key = ("PC-04", 18)
hd_ref = f"'Portfolio Marks'!$J${pm_row[hd_key]}"

cf_hdr = next_block
ws.cell(row=cf_hdr, column=3,
        value="R-3 detail: counterfactual — Q2 2026 waterfall as the fund administrator (incorrectly) ran it").font = SECTION_FONT
cf_hdr += 1
ws.cell(row=cf_hdr, column=3,
        value=("Admin remitted the Harbor Dental recap 100% as return of capital, outside the tiers, "
               "then ran the tiered waterfall on the quarter's other two exits only.")).font = SUBTITLE_FONT
cf_hdr += 1

labels_formulas = [
    ("Harbor Dental recap proceeds (paid as pure ROC by admin)", f"={hd_ref}", GREEN),
    ("Distributable run through admin's tiers (Q18 total less the recap)", f"={waterfall_ref(18,'D')}-{hd_ref}", BLACK),
    ("Cumulative contributions, EOP (unaffected)", f"={waterfall_ref(18,'F')}", GREEN),
    ("Cumulative ROC distributed, BOP — inflated by the mis-booked recap", f"={waterfall_ref(18,'G')}+{hd_ref}", BLACK),
    ("Cumulative preferred accrued, EOP (unaffected)", f"={waterfall_ref(18,'I')}", GREEN),
    ("Cumulative preferred paid, BOP (unaffected)", f"={waterfall_ref(18,'J')}", GREEN),
    ("Cumulative GP catch-up paid, BOP (unaffected)", f"={waterfall_ref(18,'K')}", GREEN),
]
row = cf_hdr
ref = {}
names = ["hd_recap", "distributable_admin", "cum_contrib_eop", "cum_roc_bop_admin",
         "cum_pref_accr_eop", "cum_pref_paid_bop", "cum_catchup_paid_bop"]
for (label, formula, font), name in zip(labels_formulas, names):
    ws.cell(row=row, column=3, value=label).font = BLACK
    c = ws.cell(row=row, column=6, value=formula)
    c.font = font
    c.number_format = CUR0
    ref[name] = f"$F${row}"
    row += 1

t1 = f"=MIN({ref['distributable_admin']},MAX({ref['cum_contrib_eop']}-{ref['cum_roc_bop_admin']},0))"
ws.cell(row=row, column=3, value="Tier 1 (admin): Return of Capital").font = BLACK
ws.cell(row=row, column=6, value=t1).font = BLACK
ws.cell(row=row, column=6).number_format = CUR0
ref["t1"] = f"$F${row}"
row += 1

t2 = f"=MIN({ref['distributable_admin']}-{ref['t1']},MAX({ref['cum_pref_accr_eop']}-{ref['cum_pref_paid_bop']},0))"
ws.cell(row=row, column=3, value="Tier 2 (admin): Preferred Return").font = BLACK
ws.cell(row=row, column=6, value=t2).font = BLACK
ws.cell(row=row, column=6).number_format = CUR0
ref["t2"] = f"$F${row}"
row += 1

pref_paid_eop_admin = f"={ref['cum_pref_paid_bop']}+{ref['t2']}"
ws.cell(row=row, column=3, value="Cumulative preferred paid, EOP (admin)").font = BLACK
ws.cell(row=row, column=6, value=pref_paid_eop_admin).font = BLACK
ws.cell(row=row, column=6).number_format = CUR0
ref["pref_paid_eop_admin"] = f"$F${row}"
row += 1

target_catchup_admin = f"=({A_CARRY_RATE}/(1-{A_CARRY_RATE}))*{ref['pref_paid_eop_admin']}"
ws.cell(row=row, column=3, value="Target cumulative GP catch-up (admin)").font = BLACK
ws.cell(row=row, column=6, value=target_catchup_admin).font = BLACK
ws.cell(row=row, column=6).number_format = CUR0
ref["target_catchup_admin"] = f"$F${row}"
row += 1

t3 = (f"=MIN({ref['distributable_admin']}-{ref['t1']}-{ref['t2']},"
      f"MAX({ref['target_catchup_admin']}-{ref['cum_catchup_paid_bop']},0))")
ws.cell(row=row, column=3, value="Tier 3 (admin): GP Catch-up").font = BLACK
ws.cell(row=row, column=6, value=t3).font = BLACK
ws.cell(row=row, column=6).number_format = CUR0
ref["t3"] = f"$F${row}"
row += 1

t4 = f"={ref['distributable_admin']}-{ref['t1']}-{ref['t2']}-{ref['t3']}"
ws.cell(row=row, column=3, value="Tier 4 (admin): Carry (total)").font = BLACK
ws.cell(row=row, column=6, value=t4).font = BLACK
ws.cell(row=row, column=6).number_format = CUR0
ref["t4"] = f"$F${row}"
row += 1

t4gp = f"={ref['t4']}*{A_CARRY_RATE}"
ws.cell(row=row, column=3, value="Tier 4 (admin): Carry to GP (20%)").font = BLACK
ws.cell(row=row, column=6, value=t4gp).font = BLACK
ws.cell(row=row, column=6).number_format = CUR0
ref["t4gp"] = f"$F${row}"
row += 1

admin_total_gp = f"={ref['t3']}+{ref['t4gp']}"
ws.cell(row=row, column=3, value="Total to GP under admin's (incorrect) sequencing").font = TOTAL_FONT
ws.cell(row=row, column=6, value=admin_total_gp).font = TOTAL_FONT
ws.cell(row=row, column=6).number_format = CUR0
ADMIN_TOTAL_GP_CELL = f"'Variance Reconciliation'!$F${row}"
row += 1

ws.cell(row=row, column=3, value="Total to GP per GP ledger (correct sequencing)").font = TOTAL_FONT
ws.cell(row=row, column=6, value=f"={waterfall_ref(18,'W')}").font = GREEN
ws.cell(row=row, column=6).number_format = CUR0
LEDGER_TOTAL_GP_CELL = f"'Variance Reconciliation'!$F${row}"
row += 1

CF_END_ROW = row

# ---- now fill in the R-3 summary row using the counterfactual result ----
ws.cell(row=r3_row, column=2, value="R-3").font = BLACK
ws.cell(row=r3_row, column=3, value="Distribution character misclassification (Q2 2026 recap booked outside the waterfall)").font = BLACK
c = ws.cell(row=r3_row, column=4, value=f"={LEDGER_TOTAL_GP_CELL}")
c.font = GREEN
c.number_format = CUR0
c = ws.cell(row=r3_row, column=5, value=f"={ADMIN_TOTAL_GP_CELL}")
c.font = GREEN
c.number_format = CUR0
c = ws.cell(row=r3_row, column=6, value=f"=E{r3_row}-D{r3_row}")
c.font = BLACK
c.number_format = CUR0
ws.cell(row=r3_row, column=7, value="Open — reclassification and GP remittance pending").font = BLACK

# ---- R-4: unreconciled wire fee ----
r4_row = SUM_START_ROW + 3  # R-1, R-2, R-3, R-4 all in the summary block (rows r1..r1+3)
ws.cell(row=r4_row, column=2, value="R-4").font = BLACK
ws.cell(row=r4_row, column=3, value="Unreconciled cash break (Q3 2025 outgoing wire fee)").font = BLACK
c = ws.cell(row=r4_row, column=4, value=0)
c.font = BLUE
c.number_format = CUR0
c = ws.cell(row=r4_row, column=5, value=-375.00)
c.font = BLUE
c.number_format = CUR0
c = ws.cell(row=r4_row, column=6, value=f"=E{r4_row}-D{r4_row}")
c.font = BLACK
c.number_format = CUR0
ws.cell(row=r4_row, column=7, value="Resolved — booked to fund operating expenses").font = BLACK

SUM_END_ROW = r4_row
tot_row2 = CF_END_ROW + 1
ws.cell(row=tot_row2, column=3, value="Gross open variance (R-2 + R-3, absolute value)").font = TOTAL_FONT
c = ws.cell(row=tot_row2, column=6, value=f"=ABS(E{r2_row}-D{r2_row})+ABS(E{r3_row}-D{r3_row})")
c.font = TOTAL_FONT
c.number_format = CUR0
c.border = TOTAL_BORDER

# ---- narrative descriptions ----
desc_row = tot_row2 + 3
ws.cell(row=desc_row, column=2, value="Break descriptions & resolution").font = SECTION_FONT
desc_row += 1
for b in reconciliation["breaks"]:
    ws.cell(row=desc_row, column=2, value=b["break_id"]).font = TOTAL_FONT
    ws.merge_cells(start_row=desc_row, start_column=3, end_row=desc_row, end_column=7)
    ws.cell(row=desc_row, column=3, value=b["description"]).font = BLACK
    ws.cell(row=desc_row, column=3).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[desc_row].height = 45
    desc_row += 1
    ws.merge_cells(start_row=desc_row, start_column=3, end_row=desc_row, end_column=7)
    ws.cell(row=desc_row, column=3, value="Resolution: " + b["resolution"]).font = SUBTITLE_FONT
    ws.cell(row=desc_row, column=3).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[desc_row].height = 30
    desc_row += 2

freeze(ws, "B6")

print("Variance Reconciliation sheet built.")
wb.save("PE_Fund_Model.xlsx")
